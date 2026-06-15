#!/usr/bin/env python3
"""Deterministic extractor: converts the source workbook into reviewable JSON
seed files (reference.json + orders.json). Run from the workbook directory:

    python3 backend/prisma/seed-data/extract.py "UFP Order Tracker (1).xlsx"

The output JSON is committed and consumed by prisma/seed; the workbook itself
is NOT needed at runtime. Re-run only when the source spreadsheet changes.
"""
import json
import sys
import datetime
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

SRC = sys.argv[1] if len(sys.argv) > 1 else "UFP Order Tracker (1).xlsx"
OUT_DIR = Path(__file__).resolve().parent

STANDARD_COLOR_NAMES = {
    "GLOSSY BLACK",
    "GLOSSY WHITE",
    "SILVER FROST",
    "CHARCOAL METALLIC",
    "VICTORY RED",
    "INDIGO BLUE",
}


def cell(v):
    """Normalize a cell value: dates -> YYYY-MM-DD, blanks -> None."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    return v


def as_bool(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("yes", "true", "y"):
        return True
    if s in ("no", "false", "n"):
        return False
    return None


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def as_int(v):
    n = num(v)
    return int(round(n)) if n is not None else None


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    # ---------- REFERENCE / BASE DATA ----------
    md = wb["Master Data"]

    # Process stages (OTC) — B/C columns rows 4..13
    stages = []
    for r in range(4, 14):
        order = md.cell(r, 2).value
        name = cell(md.cell(r, 3).value)
        if order is not None and name:
            stages.append({"order": int(order), "name": name})

    # Departure/destination ports with sailing times — E..I rows 6..9
    ports = []
    for r in range(6, 10):
        name = cell(md.cell(r, 6).value)  # F
        if not name:
            continue
        ports.append({
            "name": name,
            "sailingDays": as_int(md.cell(r, 7).value),   # G
            "freight": num(md.cell(r, 8).value),          # H
            "inland": num(md.cell(r, 9).value),           # I
        })
    origin_port = cell(md.cell(5, 7).value)  # G5 = Jebel Ali

    # Stocking locations -> arrival port — K/L/M rows 5..15
    locations = []
    for r in range(5, 16):
        name = cell(md.cell(r, 12).value)   # L
        if not name:
            continue
        locations.append({
            "name": name,
            "arrivalPort": cell(md.cell(r, 13).value),  # M
        })

    # Shipping lines — E/F/G rows 12..17
    shipping_lines = []
    for r in range(12, 18):
        name = cell(md.cell(r, 6).value)  # F
        if not name:
            continue
        shipping_lines.append({
            "name": name,
            "trackingUrl": cell(md.cell(r, 7).value),  # G
        })

    # Global constants
    config = {
        "sheetsPerSkid": as_int(md.cell(4, 16).value),       # P4 = 200
        "downpaymentPct": num(md.cell(12, 16).value),        # P12 = 0.5
        "containerMaxM2": num(md.cell(14, 16).value),        # P14 = 8600
        "leadTimeStandard": as_int(md.cell(4, 19).value),    # S4 = 45
        "leadTimeNonStandard": as_int(md.cell(5, 19).value), # S5 = 90
        "originPort": origin_port,
    }

    # Product catalog (Pricing Table)
    pt = wb["Pricing Table - eff 27-Jan-2026"]
    pricing_note = cell(pt.cell(1, 20).value)  # T1
    products = []
    colors = {}
    for r in range(3, pt.max_row + 1):
        code1 = pt.cell(r, 1).value
        if code1 is None:
            continue
        if str(code1).strip().upper().startswith("NOTE"):
            break
        name = cell(pt.cell(r, 12).value)       # L Color
        vendor = cell(pt.cell(r, 13).value)     # M Vendor Color Code
        products.append({
            "partNo": str(code1).strip(),
            "custPartNo": cell(pt.cell(r, 2).value),
            "itemType": cell(pt.cell(r, 3).value),
            "surface": cell(pt.cell(r, 4).value),
            "construction": cell(pt.cell(r, 5).value),
            "thickness": cell(pt.cell(r, 6).value),
            "widthIn": num(pt.cell(r, 7).value),
            "widthMm": num(pt.cell(r, 8).value),
            "lengthIn": num(pt.cell(r, 9).value),
            "lengthMm": num(pt.cell(r, 10).value),
            "description": cell(pt.cell(r, 11).value),
            "colorName": name,
            "vendorColorCode": vendor,
            "pricePerSqft": num(pt.cell(r, 14).value),
            "pricePerM2": num(pt.cell(r, 15).value),
            "pricePerMsq": num(pt.cell(r, 16).value),
            "pricePerSheet": num(pt.cell(r, 17).value),
            "leadTimeDays": as_int(pt.cell(r, 18).value),
        })
        if vendor and vendor not in colors:
            colors[vendor] = {
                "code": vendor,
                "name": name,
                "isStandard": bool(name and name.upper() in STANDARD_COLOR_NAMES),
            }

    reference = {
        "stages": stages,
        "ports": ports,
        "stockingLocations": locations,
        "shippingLines": shipping_lines,
        "colors": list(colors.values()),
        "products": products,
        "config": config,
        "pricingNote": pricing_note,
    }

    # ---------- TRANSACTIONAL DATA ----------
    os_ws = wb["Order Summary"]
    # column index (1-based) -> field name
    SUMMARY_MAP = {
        1: "siNo", 2: "poNo", 3: "rev", 4: "concat", 5: "status",
        6: "poDate", 7: "active", 8: "skids", 9: "stockingLocation",
        10: "portOfDest", 11: "poValue", 12: "totalM2", 13: "piNo",
        14: "piDate", 15: "poToPi", 16: "piValue", 17: "dpDate",
        18: "piToDp", 19: "dpAmount", 20: "productionEtc", 21: "shippingEta",
        22: "bol", 23: "isf", 24: "containerNo", 25: "shippingLine",
        26: "shippingUrl", 27: "actualDeparture", 28: "dpToShip", 29: "ciNo",
        30: "ciDate", 31: "revisionSent", 32: "freight", 33: "inland",
        34: "ciValue", 35: "balanceDue", 36: "bpDate", 37: "ciToBp",
        38: "bpAmount", 39: "telexDate", 40: "bpToTelex", 41: "arrivalDate",
    }
    INT_FIELDS = {"siNo", "rev", "poToPi", "piToDp", "dpToShip", "ciToBp", "bpToTelex"}
    FLOAT_FIELDS = {"skids", "poValue", "totalM2", "piValue", "dpAmount",
                    "freight", "inland", "ciValue", "balanceDue", "bpAmount"}
    # date/string fields that may also hold text like "COMPLETE"/"N/A"
    STR_FIELDS = set(SUMMARY_MAP.values()) - INT_FIELDS - FLOAT_FIELDS - {"active"}

    orders = {}
    for r in range(4, os_ws.max_row + 1):
        po_no = os_ws.cell(r, 2).value
        if po_no is None or str(po_no).strip() == "":
            continue
        rec = {}
        for col, field in SUMMARY_MAP.items():
            raw = os_ws.cell(r, col).value
            if field == "active":
                rec[field] = as_bool(raw)
            elif field in INT_FIELDS:
                rec[field] = as_int(raw)
            elif field in FLOAT_FIELDS:
                rec[field] = num(raw)
            else:
                rec[field] = cell(raw)
        rec["poNo"] = str(rec["poNo"]).strip()
        rec["lines"] = []
        key = (rec["poNo"], rec.get("rev") or 0)
        orders[key] = rec

    # Order Details -> lines
    od = wb["Order Details"]
    LINE_MAP = {
        8: "lineNo", 9: "partNo", 10: "custPartNo", 11: "size",
        12: "widthMm", 13: "lengthMm", 14: "color", 15: "qtyMsf",
        16: "qtyM2", 17: "sheets", 18: "skids", 19: "unitMsf",
        20: "unitM2", 21: "extPo", 22: "extInv", 23: "leadTime", 24: "notes",
    }
    L_INT = {"lineNo", "leadTime"}
    L_FLOAT = {"widthMm", "lengthMm", "qtyMsf", "qtyM2", "sheets", "skids",
               "unitMsf", "unitM2", "extPo", "extInv"}
    orphan_lines = 0
    for r in range(2, od.max_row + 1):
        po_no = od.cell(r, 1).value
        if po_no is None or str(po_no).strip() == "":
            continue
        rev = as_int(od.cell(r, 2).value) or 0
        line = {}
        for col, field in LINE_MAP.items():
            raw = od.cell(r, col).value
            if field in L_INT:
                line[field] = as_int(raw)
            elif field in L_FLOAT:
                line[field] = num(raw)
            else:
                line[field] = cell(raw)
        key = (str(po_no).strip(), rev)
        if key in orders:
            orders[key]["lines"].append(line)
        else:
            orphan_lines += 1
            # keep orphan lines under a synthesized header so nothing is lost
            orders[key] = {
                "poNo": str(po_no).strip(),
                "rev": rev,
                "status": cell(od.cell(r, 3).value),
                "poDate": cell(od.cell(r, 4).value),
                "active": as_bool(od.cell(r, 5).value),
                "stockingLocation": cell(od.cell(r, 6).value),
                "shippingEta": cell(od.cell(r, 7).value),
                "lines": [line],
                "_synthesized": True,
            }

    orders_list = list(orders.values())

    (OUT_DIR / "reference.json").write_text(json.dumps(reference, indent=2))
    (OUT_DIR / "orders.json").write_text(json.dumps(orders_list, indent=2))

    # summary
    print("REFERENCE:")
    print(f"  stages={len(stages)} ports={len(ports)} locations={len(locations)}"
          f" shippingLines={len(shipping_lines)} colors={len(colors)} products={len(products)}")
    print(f"  config={config}")
    print("TRANSACTIONAL:")
    print(f"  orders={len(orders_list)} (synthesized headers from orphan lines: {orphan_lines})")
    print(f"  total lines={sum(len(o['lines']) for o in orders_list)}")
    print(f"Wrote {OUT_DIR/'reference.json'} and {OUT_DIR/'orders.json'}")


if __name__ == "__main__":
    main()
