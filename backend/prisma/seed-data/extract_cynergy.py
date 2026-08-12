#!/usr/bin/env python3
"""Deterministic extractor for the Cynergy workbook.

    python3 backend/prisma/seed-data/extract_cynergy.py "docs for llm reference /Cynergy Order Tracker.xlsx"

Writes:
  * cynergy-reference.json  — master data, colours, products, color-stock matrix
  * cynergy-orders.json     — Order Summary + Order Details

Cynergy's workbook differs from UFP's in ways that matter here:

  * The price sheet is called "Price Sheet", headers sit on row 2 with a units row
    on row 3, and data starts on row 4 (UFP: headers row 1, data row 3).
  * Column A is the Full Item Description and is the catalogue key -- only ~1/3 of
    items carry a product code, and the workbook's own VLOOKUPs all target column A.
  * Price per m2 is the entered figure; $/sqft, $/MSQ and $/sheet are derived from it.
  * Master Data lives in slightly different cells (container max is P15, not P14) and
    adds a "Final Payment (days)" constant.
  * Order Summary has fewer columns than UFP (no PI/DP block); Order Details keys
    lines by Full Item Description and prices by sheet, not MSF.

Output JSON is committed and consumed by prisma/seed; the workbook is not needed at
runtime. Re-run only when the source spreadsheet changes.
"""
import json
import sys
import datetime
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

SRC = sys.argv[1] if len(sys.argv) > 1 else "Cynergy Order Tracker.xlsx"
OUT_DIR = Path(__file__).resolve().parent

# Cynergy derives $/sqft from $/m2 with this divisor (Price Sheet column O).
SQFT_PER_M2 = 10.765


def cell(v):
    """Normalize a cell value: dates -> YYYY-MM-DD, blanks -> None."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    return v


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def as_int(v):
    n = num(v)
    return int(round(n)) if n is not None else None


def as_bool(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("yes", "true", "y"):
        return True
    if s in ("no", "false", "n"):
        return False
    return None


def extract_color_stock(ws):
    """Color Sheet: True/False matrix of which W×L sizes each colour stocks.

    Layout: row 4 labels, row 5 size headers (lengths D–O, widths Q–U),
    data from row 6. Door-panel sizes live in W–Y as a separate list.
    """
    length_headers = []
    for c in range(4, 16):  # D..O
        h = cell(ws.cell(5, c).value)
        if h is not None:
            length_headers.append({"col": c, "label": str(h).rstrip('"')})

    width_headers = []
    for c in range(17, 22):  # Q..U
        h = cell(ws.cell(5, c).value)
        if h is not None:
            width_headers.append({"col": c, "label": str(h).rstrip('"')})

    door_panels = []
    for r in range(5, 20):
        w = num(ws.cell(r, 24).value)  # X
        l = num(ws.cell(r, 25).value)  # Y
        if w is None or l is None:
            continue
        door_panels.append({"widthIn": w, "lengthIn": l})

    colors = []
    for r in range(6, ws.max_row + 1):
        code = cell(ws.cell(r, 2).value)  # B
        name = cell(ws.cell(r, 3).value)  # C
        if not code and not name:
            continue
        lengths = []
        for h in length_headers:
            if ws.cell(r, h["col"]).value is True:
                lengths.append(h["label"])
        widths = []
        for h in width_headers:
            if ws.cell(r, h["col"]).value is True:
                widths.append(h["label"])
        colors.append({
            "code": code,
            "name": name,
            "lengths": lengths,
            "widths": widths,
        })

    return {
        "lengths": [h["label"] for h in length_headers],
        "widths": [h["label"] for h in width_headers],
        "doorPanels": door_panels,
        "colors": colors,
    }


def extract_orders(wb):
    """Order Summary (header row 2, data from row 3) + Order Details (from row 2)."""
    os_ws = wb["Order Summary"]
    # Cynergy's summary is shorter than UFP's — no PI / DP / balance block.
    SUMMARY_MAP = {
        1: "siNo", 2: "poNo", 3: "rev", 4: "concat", 5: "status",
        6: "poDate", 7: "active", 8: "skids", 9: "stockingLocation",
        10: "portOfDest", 11: "poValue", 12: "totalM2",
        13: "productionEtc", 14: "shippingEta", 15: "bol", 16: "isf",
        17: "containerNo", 18: "shippingLine", 19: "shippingUrl",
        20: "actualDeparture", 21: "ciNo", 22: "ciDate", 23: "freight",
        24: "ciValue", 25: "telexDate",
    }
    INT_FIELDS = {"siNo", "rev"}
    FLOAT_FIELDS = {"skids", "poValue", "totalM2", "freight", "ciValue"}

    orders = {}
    for r in range(3, os_ws.max_row + 1):
        po_no = os_ws.cell(r, 2).value
        if po_no is None or str(po_no).strip() == "":
            continue
        rec = {}
        for col, field in SUMMARY_MAP.items():
            raw = os_ws.cell(r, col).value
            if field == "active":
                rec[field] = as_bool(raw)
            elif field in INT_FIELDS:
                rec[field] = as_int(raw)
            elif field in FLOAT_FIELDS:
                rec[field] = num(raw)
            else:
                rec[field] = cell(raw)
        rec["poNo"] = str(rec["poNo"]).strip()
        # Fields UFP has that Cynergy's sheet does not — keep keys for a uniform seed.
        for missing in (
            "piNo", "piDate", "poToPi", "piValue", "dpDate", "piToDp", "dpAmount",
            "dpToShip", "revisionSent", "inland", "balanceDue", "bpDate", "ciToBp",
            "bpAmount", "bpToTelex", "arrivalDate",
        ):
            rec.setdefault(missing, None)
        # Default inland from Master Data (Savannah) when freight is present.
        if rec.get("freight") is not None and rec.get("inland") is None:
            rec["inland"] = 1000
        rec["lines"] = []
        key = (rec["poNo"], rec.get("rev") or 0)
        orders[key] = rec

    od = wb["Order Details"]
    # Lines are keyed by Full Item Description (catalogue partNo).
    orphan_lines = 0
    for r in range(2, od.max_row + 1):
        po_no = od.cell(r, 1).value
        if po_no is None or str(po_no).strip() == "":
            continue
        rev = as_int(od.cell(r, 2).value) or 0
        width_in = num(od.cell(r, 7).value)
        length_in = num(od.cell(r, 8).value)
        color = cell(od.cell(r, 9).value)
        part_no = cell(od.cell(r, 10).value)  # Full Item Description
        width_mm = num(od.cell(r, 11).value)
        length_mm = num(od.cell(r, 12).value)
        sheets = num(od.cell(r, 13).value)
        qty_m2 = num(od.cell(r, 14).value)
        skids = num(od.cell(r, 15).value)
        unit_sheet = num(od.cell(r, 16).value)
        unit_m2 = num(od.cell(r, 17).value)
        ext_inv = num(od.cell(r, 18).value)
        notes = cell(od.cell(r, 19).value)
        size = None
        if width_in is not None and length_in is not None:
            size = f'{width_in}" x {length_in}"'
        line = {
            "lineNo": as_int(od.cell(r, 6).value),
            "partNo": part_no,
            "custPartNo": None,
            "size": size,
            "widthMm": width_mm,
            "lengthMm": length_mm,
            "color": color,
            "qtyMsf": None,
            "qtyM2": qty_m2,
            "sheets": sheets,
            "skids": skids,
            "unitMsf": None,  # Cynergy orders have no MSF column
            "unitSheet": unit_sheet,
            "unitM2": unit_m2,
            "extPo": ext_inv,
            "extInv": ext_inv,
            "notes": notes,
        }
        key = (str(po_no).strip(), rev)
        if key in orders:
            orders[key]["lines"].append(line)
        else:
            orphan_lines += 1
            orders[key] = {
                "poNo": str(po_no).strip(),
                "rev": rev,
                "status": cell(od.cell(r, 3).value),
                "active": as_bool(od.cell(r, 4).value),
                "stockingLocation": cell(od.cell(r, 5).value),
                "lines": [line],
                "_synthesized": True,
            }

    return list(orders.values()), orphan_lines


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    md = wb["Master Data"]

    # Process stages (OTC) -- B/C rows 4..13. Cynergy has no Proforma Invoice or
    # Downpayment stage and adds "Arrived at Destination Port".
    stages = []
    for r in range(4, 14):
        order = md.cell(r, 2).value
        name = cell(md.cell(r, 3).value)
        if order is not None and name:
            stages.append({"order": int(order), "name": name})

    # Destination ports with sailing times -- E..I rows 6..9 (Cynergy ships to Savannah only)
    ports = []
    for r in range(6, 10):
        name = cell(md.cell(r, 6).value)  # F
        if not name:
            continue
        ports.append({
            "name": name,
            "sailingDays": as_int(md.cell(r, 7).value),   # G
            "freight": num(md.cell(r, 8).value),          # H
            "inland": num(md.cell(r, 9).value),           # I
        })
    origin_port = cell(md.cell(5, 7).value)  # G5 = Jebel Ali

    # Stocking locations -> arrival port -- K/L/M rows 5..15
    locations = []
    for r in range(5, 16):
        name = cell(md.cell(r, 12).value)   # L
        if not name:
            continue
        locations.append({
            "name": name,
            "arrivalPort": cell(md.cell(r, 13).value),  # M
        })

    # Shipping lines -- E/F/G rows 12..17
    shipping_lines = []
    for r in range(12, 18):
        name = cell(md.cell(r, 6).value)  # F
        if not name:
            continue
        shipping_lines.append({
            "name": name,
            "trackingUrl": cell(md.cell(r, 7).value),  # G
        })

    config = {
        "sheetsPerSkid": as_int(md.cell(4, 16).value),        # P4  = 200
        "downpaymentPct": num(md.cell(12, 16).value),         # P12 = 0 (no downpayment)
        "finalPaymentDays": as_int(md.cell(13, 16).value),    # P13 = 10
        "containerMaxM2": num(md.cell(15, 16).value),         # P15 = 8600
        "leadTimeStandard": as_int(md.cell(4, 19).value),     # S4  = 90
        "leadTimeNonStandard": as_int(md.cell(5, 19).value),  # S5  = 90
        "originPort": origin_port,
    }

    # Colour master -- U..Y rows 5..27. Cynergy orders by its own short colour name
    # ("BLACK"), which maps to a vendor name ("GLOSSY BLACK") and code ("ASC 0096").
    color_rows = []
    for r in range(5, md.max_row + 1):
        short = cell(md.cell(r, 22).value)   # V Cynergy Color
        if not short:
            continue
        color_rows.append({
            "shortName": short,
            "name": cell(md.cell(r, 23).value),          # W Color
            "code": cell(md.cell(r, 24).value) or short,  # X Vendor Color Code (blank for MILL FINISH)
            "construction": cell(md.cell(r, 25).value),  # Y Construction
        })

    # ---------- Price Sheet ----------
    ps = wb["Price Sheet"]
    products = []
    # Keyed on the Cynergy colour name, which every row has -- MILL FINISH carries no
    # vendor colour code, so a code-keyed map would miss it.
    m2_by_color = {}
    for r in range(4, ps.max_row + 1):
        desc = cell(ps.cell(r, 1).value)  # A Full Item Description = catalogue key
        if not desc:
            continue
        # Footer notes share column A; real items always start with a thickness.
        if not str(desc)[:1].isdigit():
            continue

        width_mm = num(ps.cell(r, 8).value)
        length_mm = num(ps.cell(r, 10).value)
        price_m2 = num(ps.cell(r, 16).value)
        vendor = cell(ps.cell(r, 13).value)

        price_sqft = num(ps.cell(r, 15).value)
        price_msq = num(ps.cell(r, 17).value)
        price_sheet = num(ps.cell(r, 18).value)
        # Recover any column whose cached formula result is missing.
        if price_m2 is not None:
            if price_sqft is None:
                price_sqft = price_m2 / SQFT_PER_M2
            if price_msq is None:
                price_msq = (price_m2 / SQFT_PER_M2) * 1000
            if price_sheet is None and width_mm is not None and length_mm is not None:
                price_sheet = (width_mm / 1000) * (length_mm / 1000) * price_m2

        short_color = cell(ps.cell(r, 14).value)  # N Cynergy Color
        code1 = ps.cell(r, 2).value  # B Product Code 1
        products.append({
            "partNo": desc,
            "vendorPartNo": str(code1).strip() if code1 is not None else None,
            "custPartNo": cell(ps.cell(r, 3).value),   # C Product Code 2
            "itemType": cell(ps.cell(r, 4).value),     # D
            "surface": cell(ps.cell(r, 5).value),      # E
            "thickness": cell(ps.cell(r, 6).value),    # F
            "widthIn": num(ps.cell(r, 7).value),       # G
            "widthMm": width_mm,                       # H
            "lengthIn": num(ps.cell(r, 9).value),      # I
            "lengthMm": length_mm,                     # J
            "construction": cell(ps.cell(r, 11).value),  # K
            "colorName": cell(ps.cell(r, 12).value),     # L
            "vendorColorCode": vendor,                   # M
            "shortColorName": short_color,
            "description": None,  # Cynergy has no separate description column
            "pricePerSqft": price_sqft,
            "pricePerM2": price_m2,
            "pricePerMsq": price_msq,
            "pricePerSheet": price_sheet,
        })

        # The colour surcharge only shows up on full 49"-wide sheets; door panels are
        # priced by size, so including them would misclassify door-panel-only colours.
        if short_color and price_m2 is not None and num(ps.cell(r, 7).value) == 49:
            m2_by_color.setdefault(short_color, set()).add(round(price_m2, 4))

    # The workbook does not label colours standard / non-standard, but it prices them in
    # two tiers on 49" sheets: a base rate and a premium rate (Charcoal Metallic, Matte
    # Black, Indigo Blue, Sand Sahara, Army Green). Treat the premium tier as
    # non-standard; colours sold only as door panels fall back to standard.
    all_tiers = sorted({p for prices in m2_by_color.values() for p in prices})
    base_tier = all_tiers[0] if all_tiers else None
    colors = []
    for c in color_rows:
        tiers = m2_by_color.get(c["shortName"], set())
        premium = bool(base_tier is not None and tiers and min(tiers) > base_tier)
        colors.append({
            "code": c["code"],
            "name": c["name"],
            "shortName": c["shortName"],
            "construction": c["construction"],
            "isStandard": not premium,
        })

    color_stock = extract_color_stock(wb["Color Sheet"])

    reference = {
        "stages": stages,
        "ports": ports,
        "stockingLocations": locations,
        "shippingLines": shipping_lines,
        "colors": colors,
        "products": products,
        "config": config,
        "pricingNote": None,
        "colorStockMatrix": color_stock,
    }

    orders_list, orphan_lines = extract_orders(wb)

    (OUT_DIR / "cynergy-reference.json").write_text(json.dumps(reference, indent=2))
    (OUT_DIR / "cynergy-orders.json").write_text(json.dumps(orders_list, indent=2))

    print("CYNERGY REFERENCE:")
    print(f"  stages={len(stages)} ports={len(ports)} locations={len(locations)}"
          f" shippingLines={len(shipping_lines)} colors={len(colors)} products={len(products)}")
    print(f"  config={config}")
    print(f"  m2 tiers={all_tiers} (base={base_tier})")
    print(f"  products with a product code={sum(1 for p in products if p['vendorPartNo'])}")
    print(f"  colorStock: {len(color_stock['colors'])} colours × "
          f"{len(color_stock['lengths'])} lengths / {len(color_stock['widths'])} widths, "
          f"{len(color_stock['doorPanels'])} door panels")
    print("CYNERGY TRANSACTIONAL:")
    print(f"  orders={len(orders_list)} (synthesized headers from orphan lines: {orphan_lines})")
    print(f"  total lines={sum(len(o['lines']) for o in orders_list)}")
    print(f"Wrote {OUT_DIR/'cynergy-reference.json'} and {OUT_DIR/'cynergy-orders.json'}")


if __name__ == "__main__":
    main()
