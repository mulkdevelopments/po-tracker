#!/usr/bin/env python3
"""Extract the UFP Production Schedule into production.json. Run from the
workbook directory:

    python3 backend/prisma/seed-data/extract_production.py "UFP  Production Schedule (2).xlsx"

Columns are resolved by header name (layout-safe across workbook revisions).
Each row maps to an existing PurchaseOrder by (poNo, rev); the seed applies the
production-side fields onto the matching order.
"""
import json
import sys
import datetime
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

SRC = sys.argv[1] if len(sys.argv) > 1 else "UFP  Production Schedule.xlsx"
OUT = Path(__file__).resolve().parent / "production.json"

# Column layouts differ between workbook revisions. Map by header name when
# possible so extractors survive column inserts/deletes.
FIELD_ALIASES = {
    "poNo": ("PO#", "PO #", "PO"),
    "rev": ("Rev#", "Rev #", "Rev"),
    "soNo": ("SO", "SO#", "SO #"),
    "standardColorsOnly": ("Standard Colors Only?", "Standard Colors Only"),
    "allMaterialAvailable": ("All Material Available",),
    "productionBegin": ("Production Begin",),
    "productionComplete": ("Production Complete",),
    "dispatchFromFactory": ("Dispatch from Factory",),
    "piSent": ("PI Sent?", "PI Sent"),
    "productionStatus": ("STATUS", "Status"),
    "productionNotes": ("Notes", "Note"),
}


def cell(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def as_int(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def header_map(ws, header_row=4):
    """Map normalized header text -> 1-based column index."""
    mapping = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, c).value
        if raw is None:
            continue
        key = str(raw).strip()
        if key:
            mapping[key] = c
    return mapping


def col(mapping, *names):
    for name in names:
        if name in mapping:
            return mapping[name]
    raise KeyError(f"Missing column(s): {names}")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Order Level"]
    headers = header_map(ws)
    cols = {field: col(headers, *aliases) for field, aliases in FIELD_ALIASES.items()}
    rows = []
    for r in range(5, ws.max_row + 1):
        po = ws.cell(r, cols["poNo"]).value
        if not po:
            continue
        rows.append({
            "poNo": str(po).strip(),
            "rev": as_int(ws.cell(r, cols["rev"]).value) or 0,
            "soNo": cell(ws.cell(r, cols["soNo"]).value),
            "standardColorsOnly": cell(ws.cell(r, cols["standardColorsOnly"]).value),
            "allMaterialAvailable": cell(ws.cell(r, cols["allMaterialAvailable"]).value),
            "productionBegin": cell(ws.cell(r, cols["productionBegin"]).value),
            "productionComplete": cell(ws.cell(r, cols["productionComplete"]).value),
            "dispatchFromFactory": cell(ws.cell(r, cols["dispatchFromFactory"]).value),
            "piSent": cell(ws.cell(r, cols["piSent"]).value),
            "productionStatus": cell(ws.cell(r, cols["productionStatus"]).value),
            "productionNotes": cell(ws.cell(r, cols["productionNotes"]).value),
        })
    OUT.write_text(json.dumps(rows, indent=2))
    print(f"Wrote {len(rows)} production rows to {OUT}")
    from collections import Counter
    print("STATUS:", dict(Counter(r["productionStatus"] for r in rows)))


if __name__ == "__main__":
    main()
