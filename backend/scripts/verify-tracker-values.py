"""Compare the stored order values against the source tracker workbook (request log #1).

    python3 scripts/verify-tracker-values.py

Reads the UFP workbook's Order Summary / Order Details sheets and reports any PO whose
header PO value, gross invoice value, total m² or skid count drifts from the file, plus any
line whose columns O-V differ.
"""
import json
import os
import subprocess
import sys
import warnings

from openpyxl import load_workbook

warnings.simplefilter("ignore")

BOOK = "../docs for llm reference /UFP Order Tracker (2).xlsx"
TOL = 0.02

here = os.path.dirname(os.path.abspath(__file__))
book = os.path.normpath(os.path.join(here, "..", BOOK))

dump = subprocess.run(
    ["node", "-e", """
const { PrismaClient } = require("@prisma/client");
require("dotenv").config();
const p = new PrismaClient();
(async () => {
  const pos = await p.purchaseOrder.findMany({
    where: { company: "UFP" },
    include: { lines: { orderBy: { lineNo: "asc" } } },
  });
  process.stdout.write(JSON.stringify(pos));
  await p.$disconnect();
})();
"""],
    cwd=os.path.join(here, ".."),
    capture_output=True,
    text=True,
)
if dump.returncode != 0:
    sys.exit(dump.stderr[-2000:])

db = {(str(p["poNo"]).strip(), p["rev"]): p for p in json.loads(dump.stdout)}

wb = load_workbook(book, data_only=True)
summary = wb["Order Summary"]
details = wb["Order Details"]


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def close(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= TOL


problems = []
checked = 0

for r in range(2, summary.max_row + 1):
    po_no = summary.cell(r, 2).value
    if po_no is None or str(po_no).strip() == "":
        continue
    rev = summary.cell(r, 3).value
    if not isinstance(rev, (int, float)):  # the sheet repeats its header row
        continue
    key = (str(po_no).strip(), int(rev))
    row = db.get(key)
    if not row:
        problems.append(f"{key[0]} rev {key[1]}: in the workbook but not in the tracker")
        continue
    checked += 1
    for label, sheet_val, stored in (
        ("PO value", num(summary.cell(r, 11).value), row["poValue"]),
        ("gross invoice value", num(summary.cell(r, 16).value), row["grossInvoiceValue"]),
        ("total m2", num(summary.cell(r, 12).value), row["totalM2"]),
        ("skids", num(summary.cell(r, 8).value), row["skids"]),
    ):
        if not close(sheet_val, stored):
            problems.append(f"{key[0]} rev {key[1]}: {label} file={sheet_val} tracker={stored}")

LINE_COLS = {
    "qtyMsf": 15,
    "qtyM2": 16,
    "sheets": 17,
    "skids": 18,
    "unitMsf": 19,
    "unitM2": 20,
    "extPo": 21,
    "extInv": 22,
}

line_checked = 0
for r in range(2, details.max_row + 1):
    po_no = details.cell(r, 1).value
    if po_no is None or str(po_no).strip() == "":
        continue
    rev = details.cell(r, 2).value
    if not isinstance(rev, (int, float)):
        continue
    key = (str(po_no).strip(), int(rev))
    row = db.get(key)
    if not row:
        continue
    line_no = details.cell(r, 8).value
    line = next((l for l in row["lines"] if l["lineNo"] == line_no), None)
    if not line:
        problems.append(f"{key[0]} rev {key[1]} line {line_no}: missing from the tracker")
        continue
    line_checked += 1
    for field, col in LINE_COLS.items():
        if not close(num(details.cell(r, col).value), line[field]):
            problems.append(
                f"{key[0]} rev {key[1]} line {line_no}: {field} "
                f"file={details.cell(r, col).value} tracker={line[field]}"
            )

for p in problems:
    print(p)
print(f"\nChecked {checked} order headers and {line_checked} lines — {len(problems)} mismatch(es).")
